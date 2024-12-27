import flet as ft
import openpyxl
from datetime import datetime, timedelta
import os
import shutil


# Função para ler dados do CID
def read_cid_data(cid_code):
    try:
        wb = openpyxl.load_workbook(
            r"C:\Users\Eloizo\Desktop\Base sistema atestados\Tabela_Cid.xlsx"
        )
        sheet = wb["CIDS"]

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] == cid_code:
                return row[1]  # Retorne a descrição da coluna B

        return "CID não encontrado"
    except Exception as e:
        print(f"Erro ao ler o arquivo CID: {e}")
        return "Erro ao ler o arquivo CID"


# Função para abrir o arquivo Excel e preencher os campos
def open_excel_file(path, matricula, fields, atestado_data_table, periodo=None):
    if not path or not matricula:
        return 0, 0

    wb = openpyxl.load_workbook(path)
    sheet = wb["Base"]

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[0]) == matricula:
            # Preenche os campos com os dados do funcionário
            fields["nome"].value = row[1]
            fields["cargo"].value = row[2]
            fields["setor"].value = row[3]
            fields["responsavel"].value = row[4]
            fields["gestor"].value = row[5]

            # Trata as datas de admissão e demissão
            if isinstance(row[6], datetime):
                fields["admissao"].value = row[6].strftime("%d/%m/%Y")
            else:
                fields["admissao"].value = row[6] if row[6] else ""

            if isinstance(row[7], datetime):
                fields["demissao"].value = row[7].strftime("%d/%m/%Y")
            else:
                fields["demissao"].value = row[7] if row[7] else ""

            # Calcula o tempo de casa
            if isinstance(row[6], datetime):
                admissao_date = row[6]
                today = datetime.today()
                delta = today - admissao_date
                years, months, days = (
                    delta.days // 365,
                    (delta.days % 365) // 30,
                    (delta.days % 365) % 30,
                )
                fields["tempo_casa"].value = f"{years} Anos {months} Meses {days} Dias"
            else:
                fields["tempo_casa"].value = ""
            break

    return get_atestado_data(path, matricula, atestado_data_table, periodo)


# Função para obter os dados dos atestados
def get_atestado_data(path, matricula, atestado_data_table, periodo=None):
    wb = openpyxl.load_workbook(path)
    sheet = wb["Atestados"]

    atestado_data_table.rows.clear()

    total_atestados = 0
    total_dias = 0

    for row in sheet.iter_rows(min_row=2, values_only=True):
        if str(row[1]) == matricula:
            data_atestado = row[15]
            data_fim_atestado = row[16]
            dias_atestado = row[17]

            # Verifica se o atestado está dentro do período especificado
            if isinstance(data_atestado, datetime):
                if periodo and periodo.isdigit():
                    hoje = datetime.now().date()
                    data_limite = hoje - timedelta(days=int(periodo))
                    if data_atestado.date() < data_limite:
                        continue

                data_atestado_str = data_atestado.strftime("%d/%m/%Y")
            else:
                data_atestado_str = str(data_atestado)

            data_fim_atestado_str = (
                data_fim_atestado.strftime("%d/%m/%Y")
                if isinstance(data_fim_atestado, datetime)
                else str(data_fim_atestado)
            )
            # Obtém os dados do CID e da instituição
            cid = row[12]  # Ajuste para a coluna correta do CID
            cid_nome = row[13]  # Ajuste para a coluna correta do CID_nome
            instituicao = row[14]  # Ajuste para a coluna correta da Instituição

            # Adiciona uma nova linha à tabela de atestados
            atestado_data_table.rows.append(
                ft.DataRow(
                    cells=[
                        ft.DataCell(ft.Text(data_atestado_str)),
                        ft.DataCell(ft.Text(data_fim_atestado_str)),
                        ft.DataCell(ft.Text(str(dias_atestado))),
                        ft.DataCell(ft.Text(cid)),  # Adiciona o CID
                        ft.DataCell(ft.Text(cid_nome)),  # Adiciona o CID
                        ft.DataCell(ft.Text(instituicao)),  # Adiciona a Instituição
                    ],
                    color=(
                        ft.colors.WHITE
                        if total_atestados % 2 == 0
                        else ft.colors.GREY_100
                    ),
                )
            )

            total_atestados += 1
            total_dias += int(dias_atestado) if dias_atestado else 0

    wb.close()
    return total_atestados, total_dias


# Função chamada quando o campo de matrícula perde o foco
def on_matricula_blur(
    e,
    path_field,
    matricula_field,
    fields,
    atestado_data_table,
    periodo_field,
    total_atestados_text,
    total_dias_text,
):
    total_atestados, total_dias = open_excel_file(
        path_field.value,
        matricula_field.value,
        fields,
        atestado_data_table,
        periodo_field.value,
    )
    total_atestados_text.value = f"Quantidade de atestados: {total_atestados}"
    total_dias_text.value = f"Total de dias: {total_dias}"
    e.page.update()


# Função chamada quando o campo de dias de atestado perde o foco
def on_dias_atestado_blur(
    e, data_atestado_field, dias_atestado_field, data_fim_atestado_field
):
    if data_atestado_field.value and dias_atestado_field.value.isdigit():
        dias = int(dias_atestado_field.value) - 1
        data_atestado = datetime.strptime(data_atestado_field.value, "%d/%m/%Y")
        data_fim = data_atestado + timedelta(days=dias)
        data_fim_atestado_field.value = data_fim.strftime("%d/%m/%Y")
        data_fim_atestado_field.update()


# Função chamada quando o campo CID é alterado
def on_cid_change(e, cid_field, cid_nome_field):
    cid_code = cid_field.value
    if cid_code:
        cid_description = read_cid_data(cid_code)
        cid_nome_field.value = cid_description
        cid_nome_field.update()


# Função para exibir um alerta
def show_alert(page, message):
    alert_dialog = ft.AlertDialog(
        title=ft.Text("Atenção"),
        content=ft.Text(message),
        actions=[
            ft.TextButton(
                "OK",
                on_click=lambda e: (
                    setattr(alert_dialog, "open", False),
                    page.update(),
                ),
            ),
        ],
    )
    page.dialog = alert_dialog
    alert_dialog.open = True
    page.update()


# Função para limpar os campos
def clear_fields(fields, atestado_fields):
    for field in fields.values():
        field.value = ""
        field.update()
    for atestado_field in atestado_fields.values():
        atestado_field.value = ""
        atestado_field.update()


# Função chamada quando um arquivo é selecionado
def pick_file_result(e, result, file_path_field):
    if result:
        file_path_field.value = result.files[0].path
        e.page.update()


def pick_files_result(
    e: ft.FilePickerResultEvent, file_picker, attached_files, page, update_callback
):
    if e.files:
        attached_files.extend(e.files)
        update_callback(page, attached_files)
    file_picker.update()


# Função para salvar o atestado
def save_attached_files(matricula, attached_files):
    base_folder = r"C:\Users\Eloizo\Desktop\Base sistema atestados\DadosArquivos"
    employee_folder = os.path.join(base_folder, matricula)

    if not os.path.exists(employee_folder):
        os.makedirs(employee_folder)

    saved_files = []
    for file in attached_files:
        file_extension = os.path.splitext(file.name)[1]
        new_filename = (
            f"{matricula}_{datetime.now().strftime('%Y%m%d%H%M%S')}{file_extension}"
        )
        destination = os.path.join(employee_folder, new_filename)
        shutil.copy(file.path, destination)
        saved_files.append(destination)

    return saved_files


def save_atestado(
    e, path_field, matricula_field, fields, atestado_fields, attached_files
):
    if not matricula_field.value:
        show_alert(e.page, "Por favor, preencha a matrícula.")
        return

    file_path = path_field.value
    if not file_path:
        show_alert(e.page, "Por favor, selecione o arquivo Excel.")
        return

    wb = openpyxl.load_workbook(file_path)
    sheet = wb["Atestados"]

    last_row = sheet.max_row + 1

    # Preenche as células com os dados do atestado
    sheet.cell(row=last_row, column=1, value=last_row - 1)
    sheet.cell(row=last_row, column=2, value=matricula_field.value)
    sheet.cell(row=last_row, column=3, value=fields["nome"].value)
    sheet.cell(row=last_row, column=4, value=fields["cargo"].value)
    sheet.cell(row=last_row, column=5, value=fields["setor"].value)
    sheet.cell(row=last_row, column=6, value=fields["responsavel"].value)
    sheet.cell(row=last_row, column=7, value=fields["gestor"].value)
    sheet.cell(row=last_row, column=8, value=fields["admissao"].value)
    sheet.cell(row=last_row, column=9, value=fields["demissao"].value)
    sheet.cell(row=last_row, column=10, value=fields["tempo_casa"].value)
    sheet.cell(row=last_row, column=11, value=atestado_fields["crm"].value)
    sheet.cell(row=last_row, column=12, value=atestado_fields["medico"].value)
    sheet.cell(row=last_row, column=13, value=atestado_fields["cid"].value)
    sheet.cell(row=last_row, column=14, value=atestado_fields["cid_nome"].value)
    sheet.cell(row=last_row, column=15, value=atestado_fields["instituicao"].value)
    sheet.cell(row=last_row, column=16, value=atestado_fields["data_atestado"].value)
    sheet.cell(
        row=last_row, column=17, value=atestado_fields["data_fim_atestado"].value
    )
    sheet.cell(row=last_row, column=18, value=atestado_fields["dias_atestado"].value)

    saved_files = []
    if attached_files:
        saved_files = save_attached_files(matricula_field.value, attached_files)
        # Adicionar os nomes dos arquivos salvos na planilha
        sheet.cell(row=last_row, column=19, value=", ".join(saved_files))

    wb.save(file_path)
    wb.close()

    show_alert(
        e.page, f"Atestado salvo com sucesso! Arquivos anexados: {len(saved_files)}"
    )
    clear_fields(fields, atestado_fields)
    attached_files.clear()
    update_attached_files_text(e.page, attached_files)


# Função para ajustar a altura das linhas da tabela
def adjust_row_height(table):
    for row in table.rows:
        for cell in row.cells:
            cell.content.style = ft.TextStyle(size=13)  # Ajusta o tamanho da fonte
        row.height = 17.25


# Função principal
def main(page: ft.Page):
    page.title = "Cadastro de Atestado Médico"
    page.window_max_width = 1100
    page.window_width = 1100
    page.window_height = 1040
    page.window_max_height = 1040

    attached_files = []

    def update_attached_files_text(page, attached_files):
        attached_files_text.value = f"Arquivos anexados: {len(attached_files)}"
        page.update()

    # Configuração do seletor de arquivos
    file_picker = ft.FilePicker(
        on_result=lambda e: pick_file_result(e, e, file_path_field)
    )
    attach_picker = ft.FilePicker(
        on_result=lambda e: pick_files_result(
            e, attach_picker, attached_files, page, update_attached_files_text
        )
    )
    page.overlay.extend([file_picker, attach_picker])

    # Campos de entrada
    file_path_field = ft.TextField(label="Caminho da Planilha Excel", width=720)
    matricula_field = ft.TextField(
        label="Matrícula",
        width=250,
        on_blur=lambda e: on_matricula_blur(
            e,
            file_path_field,
            matricula_field,
            fields,
            atestado_data_table,
            atestado_fields["Período"],
            total_atestados_text,
            total_dias_text,
        ),
    )

    # Botões
    search_button = ft.IconButton(
        icon=ft.icons.SEARCH,
        on_click=lambda e: file_picker.pick_files(allowed_extensions=["xlsx"]),
    )

    attach_button = ft.ElevatedButton(
        "Anexar Arquivos",
        icon=ft.icons.ATTACH_FILE,
        on_click=lambda _: attach_picker.pick_files(
            allow_multiple=True, allowed_extensions=["pdf", "png", "jpg", "jpeg"]
        ),
    )

    attached_files_text = ft.Text("Arquivos anexados: 0")

    save_button = ft.ElevatedButton(
        text="Salvar",
        on_click=lambda e: save_atestado(
            e, file_path_field, matricula_field, fields, atestado_fields, attached_files
        ),
    )

    # Tabela de atestados
    atestado_data_table = ft.DataTable(
        columns=[
            ft.DataColumn(
                ft.Text(
                    "Data Inicio",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Data Fim",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Dias",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "CID",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "CID Descrição",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
            ft.DataColumn(
                ft.Text(
                    "Instituição",
                    style=ft.TextStyle(
                        weight=ft.FontWeight.BOLD, color=ft.colors.BLACK
                    ),
                )
            ),
        ],
        rows=[],
        bgcolor=ft.colors.BLUE_GREY_100,  # Define a cor de fundo do cabeçalho
    )
    # Container da tabela de atestados com rolagem
    table_container = ft.Container(
        content=atestado_data_table,
        bgcolor=ft.colors.BLUE_GREY_100,
        border_radius=ft.border_radius.only(top_left=5, top_right=5),
        padding=ft.padding.only(top=10, left=10, right=10),
    )

    scrollable_table = ft.Column(
        controls=[atestado_data_table],
        height=350,  # Define a altura máxima para o contêiner da tabela
        scroll="auto",  # Habilita a rolagem automática quando necessário
    )

    # Textos informativos
    total_atestados_text = ft.Text("Quantidade de atestados: 0", weight="bold")
    total_dias_text = ft.Text("Total de dias: 0", weight="bold")

    # Campos do atestado
    atestado_fields = {
        "crm": ft.TextField(label="CRM Médico", width=250),
        "medico": ft.TextField(label="Nome Médico", width=250),
        "cid": ft.TextField(
            label="CID",
            width=100,
            on_change=lambda e: on_cid_change(
                e, atestado_fields["cid"], atestado_fields["cid_nome"]
            ),
        ),
        "cid_nome": ft.TextField(label="CID Nome", width=300, disabled=True),
        "instituicao": ft.TextField(label="Instituição", width=400),
        "Período": ft.TextField(
            label="Período Dias",
            width=90,
            on_change=lambda e: on_matricula_blur(
                e,
                file_path_field,
                matricula_field,
                fields,
                atestado_data_table,
                atestado_fields["Período"],
                total_atestados_text,
                total_dias_text,
            ),
        ),
        "data_atestado": ft.TextField(
            label="Data do Atestado",
            width=250,
        ),
        "dias_atestado": ft.TextField(
            label="Dias Atestado",
            width=100,
            on_blur=lambda e: on_dias_atestado_blur(
                e,
                atestado_fields["data_atestado"],
                atestado_fields["dias_atestado"],
                atestado_fields["data_fim_atestado"],
            ),
        ),
        "data_fim_atestado": ft.TextField(
            label="Data Fim Atestado", width=250, disabled=True
        ),
    }

    # Campos do funcionário
    fields = {
        "nome": ft.TextField(label="Nome", width=250, disabled=True),
        "cargo": ft.TextField(label="Cargo", width=250, disabled=True),
        "setor": ft.TextField(label="Setor", width=250, disabled=True),
        "responsavel": ft.TextField(label="Responsável", width=250, disabled=True),
        "gestor": ft.TextField(label="Gestor", width=250, disabled=True),
        "admissao": ft.TextField(label="Admissão", width=250, disabled=True),
        "demissao": ft.TextField(label="Demissão", width=250, disabled=True),
        "tempo_casa": ft.TextField(label="Tempo de Casa", width=250, disabled=True),
    }

    # Títulos
    title = ft.Text("Cadastro de Atestado Médico", size=18, weight="bold")
    title_row = ft.Row([title], alignment="center")
    title2 = ft.Text("Cadastro novo atestado médico", size=14, weight="bold")
    title_row2 = ft.Row([title2], alignment="center")
    title3 = ft.Text("Informações dos Atestados", size=14, weight="bold")
    title_row3 = ft.Row([title3], alignment="center")

    # Layout principal
    content = ft.Column(
        [
            title_row,
            ft.Row([file_path_field, search_button, matricula_field]),
            ft.Row(
                [
                    fields["nome"],
                    fields["cargo"],
                    fields["setor"],
                    fields["responsavel"],
                ]
            ),
            ft.Row(
                [
                    fields["gestor"],
                    fields["admissao"],
                    fields["demissao"],
                    fields["tempo_casa"],
                ]
            ),
            title_row2,
            ft.Row(
                [
                    atestado_fields["data_atestado"],
                    atestado_fields["dias_atestado"],
                    atestado_fields["data_fim_atestado"],
                    atestado_fields["instituicao"],
                ]
            ),
            ft.Row(
                [
                    atestado_fields["crm"],
                    atestado_fields["medico"],
                    atestado_fields["cid"],
                    atestado_fields["cid_nome"],
                    atestado_fields["Período"],
                ]
            ),
            ft.Row([save_button, attach_button, attached_files_text]),
            title_row3,
            scrollable_table,
            ft.Row([total_atestados_text, total_dias_text]),
        ],
        expand=True,
    )

    # Adiciona o conteúdo na página
    page.add(
        ft.Container(
            content=content,
            expand=True,
            padding=20,
        )
    )

    # Redimensionamento da página
    def page_resize(e):
        content.height = page.window_height
        scrollable_table.height = 450
        scrollable_table.width = 1100
        page.update()

    page.on_resize = page_resize
    page.update()


# Inicia a aplicação
ft.app(target=main)
